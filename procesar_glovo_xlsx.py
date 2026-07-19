"""
procesar_glovo_xlsx.py — Ingesta Glovo POR PEDIDO desde los Excel de liquidacion.
CERO OCR. Soporta los dos formatos de xlsx de Glovo (columnas en distinto orden).
    python procesar_glovo_xlsx.py "C:\\ruta\\Glovo"
"""
import os, sys, csv, re, glob
from openpyxl import load_workbook

CARPETA = sys.argv[1] if len(sys.argv) > 1 else "."

def norm(s):
    s = (s or "").strip().lower()
    for a, b in [("á","a"),("é","e"),("í","i"),("ó","o"),("ú","u")]: s = s.replace(a, b)
    return s

COL = {
 "numero de factura":"settlement_ref", "fecha de factura":"fecha_factura",
 "codigo de glovo":"codigo_glovo", "fecha de entrega":"fecha_entrega",
 "tipo de servicio":"tipo_servicio", "forma de pago":"forma_pago",
 "nombre de la tienda":"tienda", "direccion de la tienda":"direccion", "id de la tienda":"id_tienda",
 "productos":"productos", "promocion producto asumida por partner":"promo_producto",
 "promocion de oferta flash a cargo del partner":"promo_oferta_flash",
 "porcentaje de comision":"comision_pct", "total comision":"total_comision",
 "tasa de acceso a la plataforma de glovo":"tasa_acceso",
 "recargo por pedido con glovo prime":"prime", "tasa de tiempo de espera":"tasa_espera",
 "tarifas de oferta flash":"tarifas_oferta_flash", "servicio de entrega":"servicio_entrega",
 "promocion df asumida por partner":"promo_df",
 "coste de incidencias sobre productos":"coste_incidencias",
 "devoluciones de incidencias sobre productos":"devol_incidencias",
 "glovos ya remunerados":"glovos_remunerados",
 "tasa de servicio pagada en efectivo":"tasa_efectivo",
 "recargo por minimo de pedido":"recargo_minimo",
}
NUMS = ["productos","promo_producto","promo_oferta_flash","comision_pct","total_comision","tasa_acceso",
        "prime","tasa_espera","tarifas_oferta_flash","servicio_entrega","promo_df","coste_incidencias",
        "devol_incidencias","glovos_remunerados","tasa_efectivo","recargo_minimo"]

def num(v):
    if v is None or v == "": return 0.0
    try: return float(v)
    except:
        try: return float(str(v).replace(",", "."))
        except: return 0.0

def iso(v):
    if v is None or v == "": return ""
    if hasattr(v, "strftime"): return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if "/" in s:
        p = s.split("/")
        if len(p) == 3: return f"{p[2]}-{int(p[0]):02d}-{int(p[1]):02d}"
    return s.split(" ")[0]

def brand_local(tienda, direccion):
    marca = tienda or ""
    m = re.search(r"(.+?)\s*-\s*MAD\s*-", marca)
    if m: marca = m.group(1)
    marca = re.sub(r"^.*LLORENTE29\s+FOOD\s+", "", marca).strip()
    calle = (direccion or "").strip().split(",")[0].strip()
    return marca, calle

orders, settl = [], {}
files = glob.glob(os.path.join(CARPETA, "*.xlsx"))
print(f"{len(files)} xlsx en {CARPETA}")
for path in files:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header = next(it)
        idx = {}
        for i, h in enumerate(header):
            key = COL.get(norm(h))
            if key and key not in idx: idx[key] = i
        if "codigo_glovo" not in idx or "settlement_ref" not in idx:
            print(f"  aviso: cabecera inesperada en {os.path.basename(path)}, lo salto"); continue
        for row in it:
            if row is None or idx["codigo_glovo"] >= len(row): continue
            cod = row[idx["codigo_glovo"]]
            if cod in (None, ""): continue
            g = lambda k: row[idx[k]] if k in idx and idx[k] < len(row) else None
            marca, calle = brand_local(g("tienda"), g("direccion"))
            o = {"settlement_ref": str(g("settlement_ref") or "").strip(), "codigo_glovo": str(cod).strip(),
                 "fecha_factura": iso(g("fecha_factura")), "fecha_entrega": iso(g("fecha_entrega")),
                 "tipo_servicio": g("tipo_servicio"), "forma_pago": g("forma_pago"),
                 "marca": marca, "calle": calle, "id_tienda": g("id_tienda")}
            for k in NUMS: o[k] = num(g(k))
            orders.append(o)
            s = settl.setdefault(o["settlement_ref"], {"settlement_ref": o["settlement_ref"], "marca": marca,
                 "calle": calle, "fecha_factura": o["fecha_factura"], "n_pedidos": 0, **{k: 0.0 for k in NUMS}})
            s["n_pedidos"] += 1
            for k in NUMS: s[k] += o[k]
    except Exception as e:
        print(f"  ERROR en {os.path.basename(path)}: {e}")

def write(path, rows, cols):
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.DictWriter(fh, fieldnames=cols, delimiter=";"); w.writeheader(); w.writerows(rows)
if orders:
    write("glovo_orders.csv", orders, list(orders[0].keys()))
    write("glovo_settlements.csv", list(settl.values()), list(next(iter(settl.values())).keys()))
    tot = sum(abs(o["productos"]) for o in orders)
    print(f"\nHecho: {len(orders)} pedidos en {len(settl)} liquidaciones. suma |productos|: {tot:,.2f}")
else:
    print("No se extrajo ningun pedido.")
